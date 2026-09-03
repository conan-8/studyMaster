#!/usr/bin/env python3
"""
import-curated.py — human-curated xlsx -> display-ready curated records.

Input:  research/sat/curate/sat_questions.xlsx, sheet 'questions', columns
          sourceId | info | prompt | A | B | C | D | gridAnswer |
          correctAnswer | rationale | diagram | bluebook
        sourceId is the raw College Board hex id (or ssqb-<hex>).

Output: research/sat/curated/ssqb-<id>.json  (self-contained, metadata
          joined from research/sat/question-bank/ssqb-<id>.json)
        research/sat/curated-index.jsonl
        research/sat/curated-import-report.txt
        research/sat/assets/figures/ssqb-<id>.png  (copied from the harvested
          figure assets for every row with diagram=yes that has one)

Validation is hard: any row that fails a check is SKIPPED and listed in the
report — never half-imported. Cross-check columns (correctAnswer, bluebook,
diagram) are authoritative per the curation sheet; disagreements with the
harvest are logged, not failures.

Requires: python3 + openpyxl. Internal_eval only — outputs are gitignored.
"""

import json
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SAT = ROOT / 'research' / 'sat'
BANK = SAT / 'question-bank'
ASSETS = SAT / 'assets'
FIGURES = ASSETS / 'figures'
CURATED = SAT / 'curated'
WORKBOOK = SAT / 'curate' / 'sat_questions.xlsx'
REPORT = SAT / 'curated-import-report.txt'
INDEX = SAT / 'curated-index.jsonl'

EXPECTED_HEADERS = ['sourceId', 'info', 'prompt', 'A', 'B', 'C', 'D',
                    'gridAnswer', 'correctAnswer', 'rationale', 'diagram', 'bluebook']

MARKUP = [('\\(', '\\)'), ('[[', ']]'), ('**', '**')]


def markup_ok(text: str) -> list[str]:
    """Return a list of balance problems for \\(\\), [[]], ** markup."""
    problems = []
    for o, c in MARKUP:
        if text.count(o) != text.count(c):
            problems.append(f'unbalanced {o}...{c} ({text.count(o)} vs {text.count(c)})')
    return problems


def clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def norm_id(raw: str) -> str:
    s = raw.strip().lower()
    return s[5:] if s.startswith('ssqb-') else s


def main() -> int:
    if not WORKBOOK.exists():
        print(f'FATAL: {WORKBOOK} not found')
        return 1

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    if 'questions' not in wb.sheetnames:
        print(f'FATAL: sheet "questions" not found (sheets: {wb.sheetnames})')
        return 1
    ws = wb['questions']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    if headers != EXPECTED_HEADERS:
        print(f'FATAL: header mismatch.\n  expected: {EXPECTED_HEADERS}\n  got:      {headers}')
        return 1

    CURATED.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')

    bank_cache: dict[str, dict] = {}

    def bank_record(qid: str) -> dict | None:
        if qid not in bank_cache:
            p = BANK / f'ssqb-{qid}.json'
            bank_cache[qid] = json.loads(p.read_text()) if p.exists() else None
        return bank_cache[qid]

    skipped: list[tuple[str, list[str]]] = []
    flags: list[str] = []
    written: list[dict] = []
    seen: set[str] = set()
    figures_copied = 0
    figures_missing: list[str] = []

    for row_no, r in enumerate(rows[1:], start=2):
        if r is None or all(v is None for v in r):
            continue
        raw_id, info, prompt, a, b, c, d, grid, correct, rationale, diagram, bluebook = r
        if raw_id is None:
            skipped.append((f'row {row_no}', ['empty sourceId']))
            continue
        qid = norm_id(str(raw_id))
        errors: list[str] = []

        if qid in seen:
            errors.append('duplicate sourceId')
        seen.add(qid)

        src = bank_record(qid)
        if src is None:
            errors.append('ID not in harvested bank (cannot join metadata)')

        info, prompt = clean(info), clean(prompt)
        opts = [clean(x) for x in (a, b, c, d)]
        grid, correct, rationale = clean(grid), clean(correct), clean(rationale)
        diag = str(diagram).strip().lower() if diagram is not None else 'no'
        bb = str(bluebook).strip().lower() if bluebook is not None else 'no'

        if not prompt:
            errors.append('empty prompt')
        n_opts = sum(1 for o in opts if o)
        if n_opts == 4:
            qtype = 'mcq'
            if grid:
                errors.append('gridAnswer set but all four options filled')
            if correct not in ('A', 'B', 'C', 'D'):
                errors.append(f'correctAnswer {correct!r} not in A-D for an mcq')
        elif n_opts == 0:
            qtype = 'grid_in'
            if not grid:
                errors.append('no options and no gridAnswer')
            if correct is None:
                errors.append('missing correctAnswer')
        else:
            errors.append(f'partial options ({n_opts}/4 filled) — must be all four or none')

        for label, text in (('info', info), ('prompt', prompt), ('rationale', rationale),
                            ('A', opts[0]), ('B', opts[1]), ('C', opts[2]), ('D', opts[3])):
            if text:
                for p in markup_ok(text):
                    errors.append(f'{label}: {p}')

        if errors:
            skipped.append((f'ssqb-{qid}' if src is not None else qid, errors))
            continue
        assert src is not None

        # --- cross-checks: user's value wins, disagreements are logged only
        if str(src['correctAnswer']).strip() != correct:
            flags.append(f'ssqb-{qid}: correctAnswer {correct!r} != harvest {src["correctAnswer"]!r} — curated value kept')
        harvest_bb = 'yes' if src['origin'] == 'bluebook' else 'no'
        if bb != harvest_bb:
            flags.append(f'ssqb-{qid}: bluebook={bb} != harvest origin={src["origin"]} — curated value kept')
        harvest_fig = bool(src.get('stimulus', {}).get('figureAsset'))
        if diag == 'yes' and not harvest_fig:
            flags.append(f'ssqb-{qid}: diagram=yes but harvest has no figure asset — crop required')
        if diag == 'no' and harvest_fig:
            flags.append(f'ssqb-{qid}: diagram=no but harvest extracted a figure — figure dropped per curation')

        # --- figure: copy the harvested standalone PNG into assets/figures/
        diagram_path = None
        if diag == 'yes':
            src_png = ASSETS / f'ssqb-{qid}.png'
            if src_png.exists():
                shutil.copy(src_png, FIGURES / f'ssqb-{qid}.png')
                diagram_path = f'assets/figures/ssqb-{qid}.png'
                figures_copied += 1
            else:
                figures_missing.append(qid)
                flags.append(f'ssqb-{qid}: diagram=yes, asset missing — record written WITHOUT diagram, crop it into assets/figures/')

        rec = {
            'sourceId': f'ssqb-{qid}',
            'origin': 'bluebook' if bb == 'yes' else 'question_bank',
            'section': src['section'],
            'domain': src['domain'],
            'skill': src['skill'],
            'difficultyOfficial': src['difficultyOfficial'],
            'difficultyInternal': src['difficultyInternal'],
            'questionType': qtype,
            'info': info,
            'prompt': prompt,
            'options': [{'id': letter, 'text': text} for letter, text in zip('ABCD', opts) if text] if qtype == 'mcq' else [],
            'gridAnswer': grid if qtype == 'grid_in' else None,
            'correctAnswer': correct,
            'rationale': rationale,
            'diagram': diagram_path,
            'sourceUrl': src['sourceUrl'],
            'harvestedAt': src['harvestedAt'],
            'curatedAt': now,
            'allowedUses': ['internal_eval'],
        }
        out = CURATED / f'ssqb-{qid}.json'
        out.write_text(json.dumps(rec, indent=2) + '\n')
        written.append({
            'sourceId': rec['sourceId'], 'section': rec['section'], 'domain': rec['domain'],
            'skill': rec['skill'], 'difficultyOfficial': rec['difficultyOfficial'],
            'questionType': qtype, 'diagram': bool(diagram_path),
            'path': f'curated/{out.name}',
        })

    INDEX.write_text('\n'.join(json.dumps(e) for e in written) + '\n')

    lines = [
        'curated import report',
        f'workbook: {WORKBOOK.relative_to(ROOT)}',
        f'run at:   {now}',
        '',
        f'rows read:        {len(rows) - 1}',
        f'records written:  {len(written)}',
        f'figures copied:   {figures_copied}',
        f'figures missing:  {len(figures_missing)} {figures_missing}',
        f'skipped rows:     {len(skipped)}',
        f'cross-check flags: {len(flags)}',
        '',
    ]
    if skipped:
        lines += ['--- SKIPPED (fix and re-run) ---']
        lines += [f'{qid}: {"; ".join(errs)}' for qid, errs in skipped]
        lines.append('')
    if flags:
        lines += ['--- CROSS-CHECK FLAGS (user value kept) ---']
        lines += flags
        lines.append('')
    REPORT.write_text('\n'.join(lines))
    print('\n'.join(lines[:10]))
    for qid, errs in skipped:
        print(f'SKIP {qid}: {"; ".join(errs)}')
    if figures_missing:
        print(f'MISSING FIGURE CROPS: {figures_missing}')
    print(f'report: {REPORT.relative_to(ROOT)}')
    return 1 if (skipped or figures_missing) else 0


if __name__ == '__main__':
    sys.exit(main())
